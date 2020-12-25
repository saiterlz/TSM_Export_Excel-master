#!/usr/bin/python
# -*- coding: utf-8 -*-
# 本程序功能:读取由TSM插件命令/tsm scan扫描完后的AH所有商品信息,包含物品名称,最低价格,平均价格,当前拍卖量,扫描时间.等
# 通过本程序,生成一张EXCEL表格来方便进行价格走势分析.
# from win32com.client import Dispatch
# 12-21计划增加运行时,检查是否已经 打开了EXCEL文件.提前关闭
# 准备制作界面
from win32com.client import Dispatch
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import numbers  # 数据格式
from openpyxl.styles import Alignment  # 对齐方式
from openpyxl.styles import Font  # 字体
from openpyxl.styles import PatternFill  # 导入填充模块
import time
import pymysql
import os
import configparser
from tkinter import *
from tkinter.filedialog import askopenfilename
from tkinter.scrolledtext import ScrolledText
import hashlib
import time
LOG_LINE_NUM = 0
count1 = 0
count2 = 0

class GUI(Tk):
    def __init__(self,parent=None):
        super().__init__()
        self.r_value = IntVar()
        self.s_value = IntVar()
        self.set_init_window()

    # 设置窗口
    def set_init_window(self):
        self.title("TSM数据处理工具_v1.2")  # 窗口名
        # self.geometry('320x160+10+10')                         #290 160为窗口大小，+10 +10 定义窗口弹出时的默认展示位置
        self.geometry('1068x681+10+10')
        # self["bg"] = "pink"                                    #窗口背景色，其他背景色见：blog.csdn.net/chl0000/article/details/7657887
        # self.attributes("-alpha",0.9)                          #虚化，值越小虚化程度越高
        # 标签
        self.lab  = Label(self, text="File_Path")
        self.lab.grid(row=0, column=0)
        self.ent = Entry(self, width=40)
        self.ent.grid(row=0, column=1)
        self.lab = Label(self, text="类型")
        self.lab.grid(row=1, column=0)
        self.lab = Label(self, text="功能")
        self.lab.grid(row=2, column=0)
        self.log_label = Label(self, text="日志")
        self.log_label.grid(row=5, column=0)
        self.log_test_label = Label(self, text="结果")
        self.log_test_label.grid(row=5, column=12)

        # 文本框
        self.log_data_Text = ScrolledText(self, width=67, height=35)  # 日志框
        self.log_data_Text.grid(row=14, column=0, columnspan=10,sticky=S + W + E + N)

        self.log_test_Text = Text(self, width=70, height=35)  # 新增加框
        self.log_test_Text.grid(row=14, column=12, columnspan=10)
        # 按钮
        self.button1 = Button(self, text='Open', command=self.get_file_path)
        self.button1.grid(row=0, column=2)
        self.button2 = Button(self, text='submit', command=self.submit)
        self.button2.grid(row=0, column=3)
        self.str_trans_choice_1_button = Radiobutton(self, text='整体分析', variable=self.r_value, value=1,
                                                     command=self.choice_1_value)
        self.str_trans_choice_1_button.grid(row=1, column=1)
        self.str_trans_choice_2_button = Radiobutton(self, text='会长关注', variable=self.r_value, value=0,
                                                     command=self.choice_2_value)
        self.str_trans_choice_2_button.grid(row=1, column=2 )
        self.str_trans_choice_3_button = Checkbutton(self, text='写入EXCEL', command=self.myEvent1)
        self.str_trans_choice_3_button.grid(row=2,column=1)
        self.str_trans_choice_4_button = Checkbutton(self, text='进行标识', command=self.myEvent2)
        self.str_trans_choice_4_button.grid(row=2,  column=2)
        self.str_trans_to_md5_button = Button(self, text="处理数据", bg="lightblue", width=10,
                                              command=self.start_main)  # 调用内部方法  加()为直接调用
        self.str_trans_to_md5_button.grid(row=1, column=11)

    # 功能函数
    def start_main(self):
        files = self.file_path
        print(ChoiceSheetName,open_write_to_excel_button,compare_button,files,path_excel)
        main(ChoiceSheetName,open_write_to_excel_button,compare_button,files,path_excel)

    # 获取当前时间
    def get_current_time(self):
        current_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
        return current_time

    # 日志动态打印
    def write_log_to_Text(self, logmsg):
        global LOG_LINE_NUM
        current_time = self.get_current_time()
        logmsg_in = str(current_time) + " "  + str(logmsg) + "\n"  # 换行
        self.log_data_Text.insert(END, logmsg_in)
        self.log_data_Text.see(END)
        self.log_data_Text.update()
        # if LOG_LINE_NUM <= 35:
        #     self.log_data_Text.insert(END, logmsg_in)
        #     LOG_LINE_NUM = LOG_LINE_NUM + 1
        # else:
        #     self.log_data_Text.delete(1.0, 2.0)
        #     self.log_data_Text.insert(END, logmsg_in)

    def choice_1_value(self):
        global ChoiceSheetName
        print(self.r_value.get())
        ChoiceSheetName=str(self.r_value.get())
        self.write_log_to_Text('整体分析被选中')

    def choice_2_value(self):
        global ChoiceSheetName
        print(self.r_value.get())
        ChoiceSheetName = str(self.r_value.get())
        self.write_log_to_Text('会长关注被选中')

    def get_file_path(self):  #获取文件路径
        self.ent.delete(0, END)  #先清空文件名框内的内容
        self.file_name = askopenfilename(filetypes=[('All Files', 'TradeSkillMaster.lua')])  # 弹出文件复选框，选择文件,可以指定文件类型以过滤
        self.ent.insert(END, self.file_name)  # 显示文件名，用insert方法把文件名添加进去

    def submit(self):  # 点击提交的时候获取button内回调函数的变量值，这里是文件路径
        # src = self.init_data_Text.get(1.0, END).strip().replace("\n", "").encode()
        self.file_path = self.ent.get().strip().replace("\n", "").encode()
        print(self.file_path)# 用组件Entry的get获取输入框内的字符串，其在组件被销毁前就被取到
        self.write_log_to_Text('文件已选择:%s' % self.file_path)
        # self.destory()  # 中断循环，即主程序跳出无限循环mainloop()，但是这里是销毁的Frame组件，因为self指的是Frame的派生
        # root.destroy()                  #同样是跳出mainloop(),但是这里销毁的是主窗口Tk(),默认情况下它是所有tkinter 组件的父容器

    def myEvent1(self):
        global count1
        global open_write_to_excel_button
        if count1 % 2 == 0:
            count1 += 1
            self.write_log_to_Text("写入EXCEL被选中")
            open_write_to_excel_button= '1'
        else:
            count1 += 1
            self.write_log_to_Text("写入EXCEL语文被取消")
            open_write_to_excel_button= '0'

    def myEvent2(self):
        global count2
        global compare_button
        if count2 % 2 == 0:
            count2 += 1
            compare_button='1'
            self.write_log_to_Text("进行标识被选中")
        else:
            count2 += 1
            compare_button='0'
            self.write_log_to_Text("进行标识被取消")


#解决有时候EXCEL打开,无法关闭,进行强行关闭的方法来自CSDN网站
def just_open(filename):
    abs_filename= os.path.abspath(filename)
    print(abs_filename)
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(abs_filename)
    xlBook.Save()
    xlBook.Close()

#将一个name.txt文件中的ID与读取到的商品名称一一对应并写入正确商品名称
def id_to_name(filename):
    # id_name = os.path.abspath(filename)
    # print(id_name)
    ItemNames = {}
    with open(id_name,'r',encoding='utf8') as id_f:
        id_ret = id_f.readlines()
        # print(id_ret)
        for i in id_ret:
            arrStr = i.splitlines()
            # print(arrStr)
            if len(arrStr) > 0:
                for v in arrStr:
                    # print(v)
                    strI = v.split(":")
                    # print(type(strI))
                    arrI = strI
                    # print(arrI)
                    if len(arrI) == 2:
                        ItemNames[arrI[0]] = arrI[1]
            id_ret = id_f.readline()
    return ItemNames

#产生正确的时间信息
def timestamp_datetime(value):
    if type(value) != int:
        value = int(value)
    format = '%Y-%m-%d %H:%M:%S'
    # value为传入的值为时间戳(整形)，如：1332888820
    value = time.localtime(value)
    ## 经过localtime转换后变成
    ## time.struct_time(tm_year=2012, tm_mon=3, tm_mday=28, tm_hour=6, tm_min=53, tm_sec=40, tm_wday=2, tm_yday=88, tm_isdst=0)
    # 最后再经过strftime函数转换为正常日期格式。
    dt = time.strftime(format, value)
    return dt

#时间的转换
def date_style_transfomation(date, format_string1="%m-%d %H:%M:%S", format_string2="%m-%d %H-%M-%S"):
    time_array = time.strptime(date, format_string1)
    str_date = time.strftime(format_string2, time_array)
    return str_date

#将数据写入mysql的命令转换
def to_db_value(file):  # 从程序 中拿 到数据
    sql_comm_list = []
    file = files
    with open(file, encoding='utf8') as f:
        ret = f.readline()
        while ret:
            ret = f.readline()
            if sprt_word in ret:
                idxName = ret.find("internalData@csvAuctionDBScan")
                # print('idxName=', idxName)
                subName = ret[5:idxName - 1]
                if subName:
                    if ret.find("lastScan"):
                        # "f@lliance - 比格沃斯@internalData@csvAuctionDBScan" 实例
                        # 格式化数据 ，例如：itemString,minBuyout,marketValue,numAuctions,quantity,lastScan\ni:14484,69000,69000,4,4,1605895492\n
                        idxStart = ret.find("lastScan")
                        subStr = ret[idxStart + 10:len(ret) - 3]
                        arrItems = subStr.split('\\n')
                        if arrItems != 0:
                            print('have data')  # 已找到需求的数据段
                            for tmp in arrItems:
                                # print('原始数据：',tmp)
                                sql_tmp = list(tmp.split(','))
                                ItemName = sql_tmp[0].split(":")
                                sql_tmp[0] = ItemName[1]
                                sql_tmp[5] = timestamp_datetime(sql_tmp[5])  # 处理时间
                                sql_tmp.append('0')
                                # print('sql数据：', sql_tmp)
                                # sql_comm = "insert into auction_history(item_id,min_price,ave_price,auction_num,quanlity,scan_time,is_del) values (%s,%s,%s,%s,%s,str_to_date(\'%s\','%%Y-%%m-%%d %%H:%%i:%%s'),%s);" % (sql_tmp[0], sql_tmp[1], sql_tmp[2], sql_tmp[3], sql_tmp[4], sql_tmp[5], sql_tmp[6])
                                # print('SQL语句',sql_comm)
                                sql_comm_list.append(sql_tmp)
    content = tuple(sql_comm_list)  # 批量写sql语句支持元组
    return content

#将数据插件mysql数据库
def insert_to_db(file):  # 从程序 中拿 到数据
    conn = pymysql.connect("119.3.224.53", "root", "Test123abc", "wowclassic")
    cursor = conn.cursor()
    start = time.clock()
    sql_comm = "insert into auction_history(item_id,min_price,ave_price,auction_num,quanlity,scan_time,is_del) values (%s,%s,%s,%s,%s,%s,%s)"
    sql_comm_list = to_db_value(file)
    # print('insert_to_db',sql_comm_list)
    try:
        # 执行sql语句 executemany
        cursor.executemany(sql_comm, sql_comm_list)
        # 执行sql语句
        conn.commit()
    except pymysql.Error as e:
        # 发生错误时回滚
        print('执行sql出错，进行回滚', e)
        conn.rollback()
    conn.close()
    end = time.clock()
    print("executemany方法用时：", end - start, "秒")
    return print('处理写入到MYSQL')


# 给分析页添加新培加的sheet页的名字到A例第row+1行.
def add_sheet_name(workbook, dates,SheetName):
    print(workbook, dates)
    ws = workbook.get_sheet_by_name(SheetName)  # 获取sheet分析这个对象
    print(ws.title)  # 验证是否正确访问这个sheet(分析）
    ws_rows_len = ws.max_row  # 行数
    ws_cols_len = ws.max_column  # 列数
    # print("读取本表的行数 %s 和列数 %s" % (ws_rows_len, ws_cols_len))
    ws.cell(row=ws_rows_len + 1, column=1).value = dates  # 将A列的日期写入到该单元格中，单元格中的内容 是用参数传递进来
    ws_rows_curent = ws_rows_len + 1  # 定位要写入的数据为当前得到的行数加1
    for i in range(2, ws_cols_len + 1):  # 开始 遍历写入单元格公式内容 ，遍历范围了列数加1，因为for循环的机制才加1。写入的数据是从第 2列开始
        this_col_name = ws.cell(row=1, column=i).value  # 验证当前表中第一行的字段值 是否存在
        if ws.cell(row=1, column=i).value != None:  # 通过ws.cell().value函数得到该 值 ，用来判断第 一行对应字段是否为None
            # 写入公式 =VLOOKUP(B$1,INDIRECT("'"&$A4&"'!A:H"),2,0)/10000
            #       "=VLOOKUP((B$1,INDIRECT("'" + dates + "'!A:H"),2,0)/10000 "
            col_letter_str = get_column_letter(i)  # 使用get_column_letter()函数得到列对应的字母，否则为数字，无法代入公式
            print("本列的物品为:%s 在 %s 列,从 %s 行,开始写入数据..." % (this_col_name, col_letter_str, ws_rows_curent))
            indirect_str = "A" + str(ws_rows_curent)  # 拼接excel 函数 INDIRECT()中表名的内容 前后要用&$表名&
            comm_strings = '=VLOOKUP(' + col_letter_str + '$1,INDIRECT("\'"&$' + indirect_str + '&"\'!A:H"),2,0)/10000'  # 将字符串拼接成为EXCEL公式，难度 ***** 五星
            # print(comm_strings)
            ws.cell(row=ws_rows_curent, column=i).value = comm_strings  # 将拼接好的公式 写入EXCEL表
            ws.cell(row=ws_rows_curent, column=i).number_format = '0.0000'  # 设置数据格式
            ws.cell(row=ws_rows_curent, column=i).alignment = Alignment(horizontal='right',
                                                                        vertical='center')  # 设置居中对齐
        else:
            break

# 开始按列找出最小值
def get_small_value_to_color(path_excel,sheetName):
    wb = load_workbook(path_excel,data_only=True)
    ws = wb.get_sheet_by_name(sheetName)
    # 设置字体样式，设置字体为 微软雅黑，单下划线，颜色为蓝色,字体加粗
    yahei_font_u = Font(name=u'微软雅黑', underline='single', color='0000FF', bold=True)
    fille = PatternFill('solid', fgColor='c6efce')  # 设置填充颜色为 橙色
    def_fille = PatternFill('solid', fgColor='FFFFFF')  # 设置填充颜色为 白色
    print(ws.title)
    ws_rows_len = ws.max_row
    print('本 sheet 表一共有 %s 行(rows)' % ws_rows_len)
    ws_cols_len = ws.max_column
    print('本 sheet 表一共有 %s 列(columns)' % ws_cols_len)
    start_row = 4  # 定义起始行,EXCEL表中的数据列,从第4行开始
    for col in range(2, ws_cols_len + 1):  # 定位列
        temp_cell_value = float(10000000.0000)
        temp_cell_pos = []
        print('当前 是 第 %s 列.' % col)
        # col_str = get_column_letter(cols)
        # print(ws[col_str])
        for row in range(start_row, ws_rows_len + 1):  # 遍历方向是列,所以选择变更 值 为行的变化.进行循环
            # cells_value = ws.cell(row=rows, column=cols).value
            cells_value = ws.cell(row=row, column=col).value
            ws.cell(row, col).fill = def_fille  # 重置当前单元格的颜色,将以前着色的单元格恢复无底色
            ws.cell(row, col).number_format = '0.0000'  # 设置数据格式
            ws.cell(row, col).alignment = Alignment(horizontal='right')  # 设置居中对齐
            if cells_value == '#N/A' or cells_value == None:  # 判断单元格中的值 等于'#N/A ,无法使用,进行下一个循环

                print('当前 单元格的值 为:%s ,此值不可用! 当前单元格的坐标, 列为: %s -- 行为: %s' % (cells_value, col, row))
                app.write_log_to_Text('当前 单元格的值 为:%s ,此值不可用! 当前单元格的坐标, 列为: %s -- 行为: %s' % (cells_value, col, row))
                continue
            elif cells_value == '#REF!' or cells_value == 0:
                print('当前 单元格的值 为:%s ,此值不可用! 当前单元格的坐标, 列为: %s -- 行为: %s' % (cells_value, col, row))
                app.write_log_to_Text('当前 单元格的值 为:%s ,此值不可用! 当前单元格的坐标, 列为: %s -- 行为: %s' % (cells_value, col, row))
                continue
            else:
                cells_value = float(cells_value)
                print('当前 单元格的值 为:%s  , 当前单元格的坐标, 列为: %s -- 行为: %s' % (cells_value, col, row))
                app.write_log_to_Text('当前 单元格的值 为:%s  , 当前单元格的坐标, 列为: %s -- 行为: %s' % (cells_value, col, row))
                if temp_cell_value > cells_value:
                    temp_cell_value = cells_value
                    temp_cell_pos = [row, col]
                    print('进行数据比较,结果是当前单元格的值 比较小.符合要求,数据为:%s ,数据的坐标为行%s ,列 %s ' % (
                        temp_cell_value, temp_cell_pos[0], temp_cell_pos[1]))
                    app.write_log_to_Text('进行数据比较,结果是当前单元格的值 比较小.符合要求,数据为:%s ,数据的坐标为行%s ,列 %s ' % (
                        temp_cell_value, temp_cell_pos[0], temp_cell_pos[1]))
                    # ws.cell(temp_cell_pos[0], temp_cell_pos[1]).fill = fille
                    # ws.cell(row - 1, col).fill = def_fille
                elif temp_cell_value == cells_value:
                    temp_cell_pos = [row, col]
                    print('进行数据比较,结果是当前单元格的值 相等.例外,数据为:%s ,数据的坐标为行%s ,列 %s ' % (
                        temp_cell_value, temp_cell_pos[0], temp_cell_pos[1]))
                    app.write_log_to_Text('进行数据比较,结果是当前单元格的值 相等.例外,数据为:%s ,数据的坐标为行%s ,列 %s ' % (
                        temp_cell_value, temp_cell_pos[0], temp_cell_pos[1]))
                else:
                    print('进行数据比较,结果是当前 单元格的值 比较大.  不符合要求,数据为:', cells_value)
                    app.write_log_to_Text('进行数据比较,结果是当前 单元格的值 比较大.  不符合要求,数据为:%s' % cells_value)
                    pass
        # ws.cell(temp_cell_pos[0],temp_cell_pos[1]).font=yahei_font_u
        ws.cell(temp_cell_pos[0], temp_cell_pos[1]).fill = fille
    print('比较大小着色完毕!进行保存')
    wb.save(path_excel)

#将分析到的数据 写入excel表中
def write_to_excel(files,sheetName,path_excel):
    file = files
    path_excel_name = path_excel
    with open(file, encoding='utf8') as f:
        ret = f.readline()
        while ret:
            ret = f.readline()
            if sprt_word in ret:
                idxName = ret.find("internalData@csvAuctionDBScan")
                # print('idxName=', idxName)
                subName = ret[5:idxName - 1]
                if subName:
                    print('服务器文件名称为:', subName)
                    if ret.find("lastScan"):
                        # "f@lliance - 比格沃斯@internalData@csvAuctionDBScan" 实例
                        # 格式化数据 ，例如：itemString,minBuyout,marketValue,numAuctions,quantity,lastScan\ni:14484,69000,69000,4,4,1605895492\n
                        idxStart = ret.find("lastScan")
                        subStr = ret[idxStart + 10:len(ret) - 3]
                        arrItems = subStr.split('\\n')
                        if os.path.exists("%s" % path_excel_name):  #抓出一个BUG,excel的文件名与服务器名称信息混在一起了
                            wb = load_workbook("%s" % path_excel_name)
                        else:
                            wb = Workbook(data_only=True)
                        # AddSheet(fmt.Sprintf("%s", time.Now().Format("01-02 15-04-05"))
                        new_sheet_name = time.strftime("%m-%d %H-%M-%S", time.localtime())
                        ws = wb.create_sheet(new_sheet_name)
                        ws.cell(1, 1).value = u"物品名称"
                        ws.column_dimensions["B"].width = 20
                        # ws.row_dimensions[1].height = 40 #行高
                        ws.cell(1, 2).value = u"最低价格"
                        ws.column_dimensions["B"].width = 10
                        ws.cell(1, 3).value = u"平均价格"
                        ws.column_dimensions["C"].width = 10
                        ws.cell(1, 4).value = u"拍卖数量"
                        ws.column_dimensions["D"].width = 10
                        ws.cell(1, 5).value = u"物品数量"
                        ws.column_dimensions["E"].width = 10
                        ws.cell(1, 6).value = u"TSM4最后更新数据时间"
                        ws.column_dimensions["F"].width = 25
                        if arrItems != 0:
                            print('have data')  # 找到需求的数据段
                            for tmp in arrItems:
                                list_tmp = list(tmp.split(','))
                                ItemName = list_tmp[0].split(":")
                                list_tmp[0] = ItemNames[ItemName[1]]  # 处理名称
                                list_tmp[5] = timestamp_datetime(list_tmp[5])
                                ws.append(list_tmp)  # 写入数据到EXCEL
                        else:
                            print('no data ,error split data !')  # 没有找到需要数据段
                    else:
                        print('no data1')
                    add_sheet_name(wb, new_sheet_name,sheetName)
                    wb.save("%s" % path_excel_name)
    return print('处理写入到EXCEL')

# main(ChoiceSheetName,open_write_to_excel_button,compare_button,files)
def main(ChoiceSheetName,open_write_to_excel_button,compare_button,files,path_excel):
    ChoiceSheetName=str(ChoiceSheetName)
    open_write_to_excel_button=str(open_write_to_excel_button)
    compare_button=str(compare_button)
    files=files
    path_excel=path_excel
    open_to_sql_button='0'
    try:
        if ChoiceSheetName == '0':
            Analysis_Sheet ='柠檬专用'
        elif ChoiceSheetName == '1':
            Analysis_Sheet = '分析'
        else:
            print('不用写入数据库')
    except Exception as err:
        print(err)
    try:
        if open_to_sql_button != '0':
            insert_to_db(files)
        else:
            print('不用写入数据库')
    except Exception as err:
        print(err)

    try:
        if open_write_to_excel_button != "0":
            write_to_excel(files, Analysis_Sheet,path_excel)
        else:
            print('不用写入EXCEL表')
    except Exception as err:
        print(err)
    try:
        if compare_button != "0":
            just_open(path_excel)
            get_small_value_to_color(path_excel, Analysis_Sheet)
        else:
            print('不用写入EXCEL表')
    except Exception as err:
        print(err)
    print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))



if __name__ == "__main__":

    proDir = os.path.split(os.path.realpath(__file__))[0]
    # 在当前文件路径下查找.ini文件
    configPath = os.path.join(proDir, "config.ini")
    print(configPath)
    conf = configparser.ConfigParser()
    # 读取.ini文件
    conf.read(configPath, encoding="utf-8-sig")
    # Analysis_Sheet0 = conf.get('value', 'Analysis_Sheet0')
    # Analysis_Sheet1 = conf.get('value', 'Analysis_Sheet1')
    path_excel = "D:\\mystudy\\TSM_Export_Excel-master\\Alliance - 比格沃斯2.xlsx"
    sprt_word = "csvAuctionDBScan"
    # files = conf.get('path', 'files')
    id_name = "D:\\mystudy\\TSM_Export_Excel-master\\nameB.txt"
    ItemNames = id_to_name(id_name)
    open_to_sql_button = '0'
    app = GUI()
    # 设置根窗口默认属性
    app.mainloop()  # 父窗口进入事件循环，可以理解为保持窗口运行，否则界面不展示
